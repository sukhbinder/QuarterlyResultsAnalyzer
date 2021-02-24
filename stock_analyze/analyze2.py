import argparse
import os
import tempfile

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import stock_analyze.AnalyzeQuarters as sa


def get_trend_analysis(df):
    dd1 = df.fillna(0)
    a = dd1.iloc[:, 4].copy()
    for ii in range(4, 10):
        dd1.iloc[:, ii] = 100.0 * dd1.iloc[:, ii] / a
    return dd1.iloc[:, 4:].fillna(0.0).round(0)


def get_data_in_excel(fname, figures=False):
    items = ["Sales", "expense", "equity", "cash", "Investments", "Interest",
             "Profit before tax", "Dividend Amount", "opm", "npm", "d2e", "roe", "eps", ]

    df, val = sa.get_pl_bal_cash_price(fname)
    pl, bal, cash = val
    df1 = sa.add_otherdata(df)

    wb = openpyxl.Workbook()
    write_raw_data(wb, df1)
    write_skip_years(df1, items, wb)
    common_sizes_in_wb(df1, wb)

    write_trend(pl, wb, sheet_name="PL")
    write_trend(bal, wb, sheet_name="BAL")
    write_trend(cash, wb, sheet_name="CASH")

    if figures:
        figures_in_workbook(fname, wb, tempdir)

    outfile = os.path.join(os.path.dirname(
        fname), "AN_{}".format(os.path.basename(fname)))
    wb.save(outfile)


def write_trend(df, wb, sheet_name="PL"):
    dft = get_trend_analysis(df)
    ws1 = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(dft, index=True, header=True):
        ws1.append(r)


def write_raw_data(wb, df1):
    ws = wb.active
    dd = df1.fillna(0).round(1).T
    for r in dataframe_to_rows(dd, index=True, header=True):
        if not "Report Date" in r:
            ws.append(r)
    write_trend(dd, wb, sheet_name="Entire_trend")


def write_skip_years(df1, items, wb):
    dd1 = df1[items].fillna(0).round(1).T

    ws1 = wb.create_sheet("Skip_Years")
    for r in dataframe_to_rows(dd1.iloc[:, ::-3], index=True, header=True):
        ws1.append(r)


def common_sizes_in_wb(df1, wb):
    cs = sa.get_commonsize_analysis(df1)
    ws1 = wb.create_sheet("Common_sizes")
    for r in dataframe_to_rows(cs, index=True, header=True):
        ws1.append(r)


def figures_in_workbook(fname, wb, tempdir):
    with tempfile.TemporaryDirectory(prefix="work_") as tempdir:
        fig = sa.get_overall(fname)
        for i, f in enumerate(fig):
            ws1 = wb.create_sheet(str(i))
            figname = "{}.png".format(i)
            figname = os.path.join(tempdir, figname)
            f.savefig(figname, dpi=150)
            img = openpyxl.drawing.image.Image(figname)
            ws1.add_image(img)
            a = ws1.cell(1, 2, 1)


def main():
    parser = argparse.ArgumentParser(description="Analyze quarter")
    parser.add_argument("file", type=str,  help="screener file")

    args = parser.parse_args()

    assert os.path.exists(args.file)
    get_data_in_excel(args.file)
